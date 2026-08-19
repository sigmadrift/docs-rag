"""스프레드시트를 행 단위 텍스트로 변환한다.

표는 행마다 의미가 완결되므로 문장 패킹(chunking.split_text)이 아니라 행을 그대로 청크로 쓴다.
한 청크에 여러 단자의 규격이 섞이면 LLM이 다른 행의 수치를 인용할 수 있기 때문이다.

실무 엑셀은 위쪽에 문서 메타(제목·문서번호·공정명), 그 아래 다단 헤더, 그 아래 데이터가 오는
형태가 흔하다. 세 구간을 이렇게 나눈다.
  - 데이터 시작: 첫 열이 숫자인 첫 행 (순번 열이 있는 표를 가정)
  - 헤더 블록: 데이터 바로 위 셀이 속한 세로 병합의 시작 행부터 (다단 헤더는 세로로 병합된다)
  - 문서 메타: 그보다 위 전부
"""
from pathlib import Path

from openpyxl import load_workbook


def _text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float):
        # 엑셀 부동소수점 오차(1.9100000000000001) 정리
        return f"{round(value, 4):g}"
    return " ".join(str(value).split())


def _filled_grid(ws) -> tuple[list[list], dict]:
    """병합 셀의 좌상단 값을 범위 전체에 채운 격자와, 셀→병합범위 맵을 돌려준다."""
    grid = [[cell.value for cell in row] for row in ws.iter_rows()]
    spans = {}
    for merged in ws.merged_cells.ranges:
        value = grid[merged.min_row - 1][merged.min_col - 1]
        for r in range(merged.min_row - 1, merged.max_row):
            for c in range(merged.min_col - 1, merged.max_col):
                grid[r][c] = value
                spans[(r, c)] = merged
    return grid, spans


def _drop_common_prefix(prev: str, current: str) -> str:
    """앞 열 헤더와 겹치는 상위 계층을 지운다.

    'SPEC (mm) C/H 기준' 다음의 'SPEC (mm) C/H (＋) 공차'는 '(＋) 공차'로 줄어든다.
    전부 겹치면(같은 헤더) 원래 값을 그대로 쓴다.
    """
    prev_words, cur_words = prev.split(), current.split()
    i = 0
    while i < len(prev_words) and i < len(cur_words) and prev_words[i] == cur_words[i]:
        i += 1
    return " ".join(cur_words[i:]) or current


def extract_rows(path: Path) -> list[str]:
    """모든 시트의 데이터 행을 각각 한 덩어리의 텍스트로 만든다."""
    workbook = load_workbook(path, data_only=True, read_only=False)
    pieces: list[str] = []
    for ws in workbook.worksheets:
        grid, spans = _filled_grid(ws)
        if not grid:
            continue
        data_start = next((i for i, row in enumerate(grid) if isinstance(row[0], int | float)), None)
        if data_start is None:  # 순번 열이 없는 시트는 건너뛴다
            continue
        merged = spans.get((data_start - 1, 0))
        head_start = merged.min_row - 1 if merged else max(data_start - 1, 0)

        meta: list[str] = []
        for row in grid[:head_start]:
            for value in (_text(v) for v in row):
                if value and value not in meta:
                    meta.append(value)
        prefix = f"[{' | '.join(meta)}] " if meta else ""

        headers = []
        for col in range(len(grid[0])):
            seen, parts = set(), []
            for row in grid[head_start:data_start]:
                value = _text(row[col])
                if value and value not in seen:
                    seen.add(value)
                    parts.append(value)
            headers.append(" ".join(parts))

        for row in grid[data_start:]:
            if not _text(row[1] if len(row) > 1 else row[0]):
                continue  # 표 아래 여백/합계 행
            fields, previous = [], ""
            for col, value in enumerate(row):
                text = _text(value)
                if not text or not headers[col]:
                    continue
                fields.append(f"{_drop_common_prefix(previous, headers[col])}: {text}")
                previous = headers[col]
            # 순번·이름만 있는 행(목차 시트 등)은 검색 가치가 없고 실제 규격 행을 밀어낸다
            if len(fields) >= 3:
                pieces.append(prefix + " | ".join(fields))
    workbook.close()
    return pieces
