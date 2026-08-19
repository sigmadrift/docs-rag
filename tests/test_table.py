"""스프레드시트 행 변환 — 실무 엑셀 형태(문서 메타 + 다단 헤더 + 데이터)를 만들어 검증한다."""
import pytest
from openpyxl import Workbook

from app.services.table import extract_rows


@pytest.fixture
def sample_xlsx(tmp_path):
    """1~2행 문서 메타, 3~4행 다단 헤더(세로/가로 병합), 5행~ 데이터."""
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "압착 조건표"
    ws["D1"] = "KRT-00-030"
    ws.merge_cells("A1:C2")  # 문서 제목 (넓은 병합)

    ws["A3"] = "순"
    ws["B3"] = "단자명"
    ws["C3"] = "SPEC"
    ws["C4"] = "기준"
    ws["D4"] = "공차"
    ws.merge_cells("A3:A4")  # 세로 병합 → 헤더 블록 높이를 알려준다
    ws.merge_cells("B3:B4")
    ws.merge_cells("C3:D3")  # 가로 병합 → 상위 계층

    ws.append([1, "731265-3", 1.05, 0.1])
    ws.append([2, "50660-9001", 1.9100000000000001, 0.05])
    ws.append([3, "이름만"])  # 필드가 적은 행
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    return path


def test_meta_and_header_are_separated(sample_xlsx):
    rows = extract_rows(sample_xlsx)

    assert rows, "행이 하나도 나오지 않음"
    # 문서 메타는 대괄호 접두사로, 열 헤더는 필드 이름으로 들어간다
    assert rows[0].startswith("[압착 조건표 | KRT-00-030]")
    assert "단자명: 731265-3" in rows[0]


def test_multilevel_header_without_repetition(sample_xlsx):
    rows = extract_rows(sample_xlsx)

    # 상위 계층(SPEC)은 첫 열에만 붙고, 다음 열에서는 공통 부분이 생략된다
    assert "SPEC 기준: 1.05" in rows[0]
    assert "| 공차: 0.1" in rows[0]
    assert rows[0].count("SPEC") == 1


def test_float_noise_is_rounded(sample_xlsx):
    rows = extract_rows(sample_xlsx)

    # 엑셀 부동소수점 오차가 그대로 노출되지 않아야 한다
    assert "1.91" in rows[1]
    assert "1.9100000000000001" not in rows[1]


def test_sparse_row_is_dropped(sample_xlsx):
    rows = extract_rows(sample_xlsx)

    # 순번·이름만 있는 행은 검색 가치가 없어 버린다
    assert len(rows) == 2
    assert all("이름만" not in r for r in rows)


def test_sheet_without_numeric_first_column_is_skipped(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.append(["메모", "내용"])
    ws.append(["가", "나"])
    path = tmp_path / "no_index.xlsx"
    wb.save(path)

    assert extract_rows(path) == []
